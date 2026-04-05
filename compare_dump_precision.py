#!/usr/bin/env python3
"""
Dump 数据精度对比脚本 — 对比 Ascend 和 H20 的 dump 数据精度
=============================================================
按 dump_data_ascend_说明文档.md 中定义的执行顺序，逐文件比较
dump_data_ascend 和 dump_data_h20 目录中的张量数据，
使用 ULP 和 SNR 指标进行精度评估，生成 Markdown 报告、可视化图表和 Excel 表格。

输出:
  1. Markdown 报告（按 dump 顺序的完整比对结果 + 汇总 + 质量评估）
  2. 可视化图表（SNR 折线图、ULP 折线图、SNR 热力图、误差累积图）
  3. Excel 表格（按 precision_result.xlsx 模板格式）

用法:
  python compare_dump_precision.py \
      --baseline-dir ./dump_data_h20 \
      --test-dir ./dump_data_ascend \
      --output-report ./dump_precision_report.md \
      --output-charts-dir ./dump_charts \
      --output-excel ./precision_result_output.xlsx

环境要求: torch, numpy, matplotlib, openpyxl
"""

import argparse
import math
import os
from dataclasses import dataclass
from typing import List, Optional, Tuple

import torch
import numpy as np


# ===========================================================================
# Metric functions (from compare_precision.py)
# ===========================================================================

def compute_bf16_ulp(values: torch.Tensor) -> torch.Tensor:
    """Compute the ULP (Unit in Last Place) for each element based on BF16 precision.

    BF16 has 7 mantissa bits, so:
        ULP = 2^(floor(log2(|value|)) - 7)
    """
    abs_val = values.abs().float()
    abs_val_clamped = abs_val.clamp(min=2.0 ** -126)
    exponent = torch.floor(torch.log2(abs_val_clamped))
    ulp = torch.pow(2.0, exponent - 7)
    return ulp


def compute_ulp_metrics(baseline: torch.Tensor, test: torch.Tensor) -> dict:
    """Compute ULP-based precision metrics.

    When both baseline and test are zero at a position, that element
    contributes 0 ULP (perfect match). When only baseline is zero,
    ULP is computed using the test value's magnitude instead.
    """
    diff = (baseline - test).abs().float()
    # Use the larger magnitude of baseline and test for ULP reference
    ref = torch.max(baseline.abs(), test.abs())
    ulp = compute_bf16_ulp(ref)
    ulp_count = diff / ulp

    # Where both are exactly zero, set ulp_count to 0 (perfect match)
    both_zero = (baseline == 0) & (test == 0)
    ulp_count[both_zero] = 0.0

    total_elements = ulp_count.numel()
    return {
        "mean_ulp": ulp_count.mean().item(),
        "max_ulp": ulp_count.max().item(),
        "within_1ulp_pct": (ulp_count <= 1.0).sum().item() / total_elements * 100.0,
        "within_2ulp_pct": (ulp_count <= 2.0).sum().item() / total_elements * 100.0,
    }


def compute_snr_db(baseline: torch.Tensor, test: torch.Tensor) -> float:
    """Compute Signal-to-Noise Ratio in dB."""
    baseline_f = baseline.float()
    noise = baseline_f - test.float()
    signal_power = (baseline_f ** 2).sum().item()
    noise_power = (noise ** 2).sum().item()
    if noise_power == 0:
        return float('inf')
    if signal_power == 0:
        return 0.0
    return 10.0 * math.log10(signal_power / noise_power)


def compute_abs_error_metrics(baseline: torch.Tensor, test: torch.Tensor) -> dict:
    """Compute absolute error metrics."""
    diff = (baseline - test).abs().float()
    return {
        "mean_abs_err": diff.mean().item(),
        "max_abs_err": diff.max().item(),
        "min_abs_err": diff.min().item(),
    }


def compute_rel_error_metrics(baseline: torch.Tensor, test: torch.Tensor) -> dict:
    """Compute relative error metrics."""
    diff = (baseline - test).abs().float()
    ref = baseline.abs().float().clamp(min=1e-12)
    rel = diff / ref
    return {
        "mean_rel_err": rel.mean().item(),
        "max_rel_err": rel.max().item(),
        "min_rel_err": rel.min().item(),
    }


def compute_basic_stats(tensor: torch.Tensor) -> dict:
    """Compute mean, min, max of a tensor."""
    t = tensor.float()
    return {
        "mean": t.mean().item(),
        "min": t.min().item(),
        "max": t.max().item(),
    }


# ===========================================================================
# Dump 顺序定义（按说明文档的执行流程）
# ===========================================================================

# 功能阶段分类（用于汇总和图表着色）
STAGE_NAMES = {
    "input": "Layer Input",
    "after_input_layernorm": "Input LayerNorm",
    "qkv_proj_input": "Attention",
    "qkv_proj_weight": "Attention",
    "after_qkv_proj": "Attention",
    "after_attn_op": "Attention",
    "after_attn_layer": "Attention Output",
    "after_attn_add_residual": "Attention+Residual",
    "after_post_attention_layernorm": "Post-Attn LayerNorm",
    "after_ffn_gate_up_proj": "Dense FFN",
    "after_ffn_act_fn": "Dense FFN",
    "experts_input": "MoE Entry",
    "experts_e_score_correction_bias": "MoE Entry",
    "after_router": "MoE Entry",
    "grouped_topk_input": "MoE Routing",
    "grouped_topk_gating_output": "MoE Routing",
    "grouped_topk_scores": "MoE Routing",
    "grouped_topk_scores_with_bias": "MoE Routing",
    "grouped_topk_after_bias_group_scores": "MoE Routing",
    "grouped_topk_group_idx": "MoE Routing",
    "grouped_topk_tmp_scores": "MoE Routing",
    "grouped_topk_topk_weights_1": "MoE Routing",
    "grouped_topk_topk_weights_2": "MoE Routing",
    "grouped_topk_topk_weights_3": "MoE Routing",
    "fused_topk_weights": "MoE Fusion",
    "fused_topk_ids": "MoE Fusion",
    "after_experts_shared": "MoE Experts",
    "after_experts_fused": "MoE Experts",
    "after_mlp_layer": "MLP+Residual",
}

# Dense 层: 12 个文件
DENSE_LAYER_DUMPS = [
    "input",
    "after_input_layernorm",
    "qkv_proj_input",
    "qkv_proj_weight",
    "after_qkv_proj",
    "after_attn_op",
    "after_attn_layer",
    "after_attn_add_residual",
    "after_post_attention_layernorm",
    "after_ffn_gate_up_proj",
    "after_ffn_act_fn",
    "after_mlp_layer",
]

# MoE 层: 最多 27 个文件
MOE_LAYER_DUMPS = [
    "input",
    "after_input_layernorm",
    "qkv_proj_input",
    "qkv_proj_weight",
    "after_qkv_proj",
    "after_attn_op",
    "after_attn_layer",
    "after_attn_add_residual",
    "after_post_attention_layernorm",
    # MoE 入口
    "experts_input",
    "experts_e_score_correction_bias",
    "after_router",
    # 专家选择 (experts_selector)
    "grouped_topk_input",
    "grouped_topk_gating_output",
    "grouped_topk_scores",
    "grouped_topk_scores_with_bias",
    "grouped_topk_after_bias_group_scores",
    "grouped_topk_group_idx",
    "grouped_topk_tmp_scores",
    "grouped_topk_topk_weights_1",
    "grouped_topk_topk_weights_2",
    "grouped_topk_topk_weights_3",
    # fused_moe 入口
    "fused_topk_weights",
    "fused_topk_ids",
    # MoE 专家输出
    "after_experts_shared",
    "after_experts_fused",
    # 层输出
    "after_mlp_layer",
]


# ===========================================================================
# Data classes
# ===========================================================================

@dataclass
class DumpMetrics:
    """Metrics for a single dump file comparison."""
    layer_idx: int
    dump_name: str
    stage: str
    filename: str
    # Shape info
    baseline_shape: list
    test_shape: list
    # Baseline (H20) stats
    baseline_mean: float
    baseline_min: float
    baseline_max: float
    # Test (Ascend) stats
    test_mean: float
    test_min: float
    test_max: float
    # ULP
    mean_ulp: float
    max_ulp: float
    within_1ulp_pct: float
    within_2ulp_pct: float
    # SNR
    snr_db: float
    # Abs error
    mean_abs_err: float
    max_abs_err: float
    min_abs_err: float
    # Rel error
    mean_rel_err: float
    max_rel_err: float
    min_rel_err: float


# ===========================================================================
# Tensor loading
# ===========================================================================

def load_tensor(file_path: str) -> torch.Tensor:
    """Load a dump .bin file as a float32 tensor."""
    try:
        tensor = torch.load(file_path, map_location='cpu', weights_only=True)
        if isinstance(tensor, np.ndarray):
            tensor = torch.from_numpy(tensor)
    except Exception:
        tensor_np = np.fromfile(file_path, dtype=np.float32)
        tensor = torch.from_numpy(tensor_np)
    return tensor.float()


def align_shapes(bl: torch.Tensor, tl: torch.Tensor) -> Tuple[torch.Tensor, torch.Tensor]:
    """Handle shape mismatch by truncating to minimum dimensions."""
    if bl.shape == tl.shape:
        return bl, tl
    slices = []
    for d in range(bl.dim()):
        min_size = min(bl.shape[d], tl.shape[d])
        slices.append(slice(0, min_size))
    slices = tuple(slices)
    return bl[slices], tl[slices]


# ===========================================================================
# Core comparison logic
# ===========================================================================

def compare_dumps(
    baseline_dir: str,
    test_dir: str,
) -> List[DumpMetrics]:
    """Compare all dump files between baseline and test directories.

    Iterates layers in order: 0, 1, ..., 39
    For each layer, iterates dump points in execution order.

    Returns:
        List of DumpMetrics in dump order.
    """
    results = []

    # Determine layer range by scanning baseline directory
    layer_indices = set()
    for fname in os.listdir(baseline_dir):
        if fname.endswith('.bin') and fname.startswith('layer'):
            prefix = fname.split('_', 1)[0]  # "layer0", "layer-1", etc.
            try:
                idx = int(prefix.replace('layer', ''))
                layer_indices.add(idx)
            except ValueError:
                continue

    # Filter out layer -1 (no longer needed)
    layer_indices.discard(-1)
    layer_indices = sorted(layer_indices)

    if not layer_indices:
        raise FileNotFoundError(f"No valid layer files found in {baseline_dir}")

    print(f"Found layers: {layer_indices[0]} to {layer_indices[-1]} ({len(layer_indices)} layers)")

    for layer_idx in layer_indices:
        # Select dump list based on layer type
        test_file = os.path.join(baseline_dir, f"layer{layer_idx}_experts_input.bin")
        if os.path.exists(test_file):
            dump_list = MOE_LAYER_DUMPS
        else:
            dump_list = DENSE_LAYER_DUMPS

        for dump_name in dump_list:
            filename = f"layer{layer_idx}_{dump_name}.bin"
            bl_path = os.path.join(baseline_dir, filename)
            tl_path = os.path.join(test_dir, filename)

            if not os.path.exists(bl_path) or not os.path.exists(tl_path):
                # Some dumps are conditional, skip if missing
                continue

            try:
                bl_tensor = load_tensor(bl_path)
                tl_tensor = load_tensor(tl_path)
            except Exception as e:
                print(f"  WARNING: Failed to load {filename}: {e}")
                continue

            orig_bl_shape = list(bl_tensor.shape)
            orig_tl_shape = list(tl_tensor.shape)

            if bl_tensor.shape != tl_tensor.shape:
                print(f"  WARNING: shape mismatch {filename}: {bl_tensor.shape} vs {tl_tensor.shape}")
                bl_tensor, tl_tensor = align_shapes(bl_tensor, tl_tensor)

            # Compute metrics
            bl_stats = compute_basic_stats(bl_tensor)
            tl_stats = compute_basic_stats(tl_tensor)
            ulp = compute_ulp_metrics(bl_tensor, tl_tensor)
            snr = compute_snr_db(bl_tensor, tl_tensor)
            abs_err = compute_abs_error_metrics(bl_tensor, tl_tensor)
            rel_err = compute_rel_error_metrics(bl_tensor, tl_tensor)

            stage = STAGE_NAMES.get(dump_name, "Unknown")

            results.append(DumpMetrics(
                layer_idx=layer_idx,
                dump_name=dump_name,
                stage=stage,
                filename=filename,
                baseline_shape=orig_bl_shape,
                test_shape=orig_tl_shape,
                baseline_mean=bl_stats["mean"],
                baseline_min=bl_stats["min"],
                baseline_max=bl_stats["max"],
                test_mean=tl_stats["mean"],
                test_min=tl_stats["min"],
                test_max=tl_stats["max"],
                mean_ulp=ulp["mean_ulp"],
                max_ulp=ulp["max_ulp"],
                within_1ulp_pct=ulp["within_1ulp_pct"],
                within_2ulp_pct=ulp["within_2ulp_pct"],
                snr_db=snr,
                mean_abs_err=abs_err["mean_abs_err"],
                max_abs_err=abs_err["max_abs_err"],
                min_abs_err=abs_err["min_abs_err"],
                mean_rel_err=rel_err["mean_rel_err"],
                max_rel_err=rel_err["max_rel_err"],
                min_rel_err=rel_err["min_rel_err"],
            ))

        if layer_idx % 5 == 0:
            print(f"  Processed layer {layer_idx}")

    print(f"\nTotal comparisons: {len(results)}")
    return results


# ===========================================================================
# Report generation (Markdown)
# ===========================================================================

def _fmt(val: float, fmt: str = ".4f") -> str:
    if math.isinf(val):
        return "inf"
    if math.isnan(val):
        return "NaN"
    return f"{val:{fmt}}"


def generate_markdown_report(
    results: List[DumpMetrics],
    baseline_dir: str,
    test_dir: str,
) -> str:
    """Generate a Markdown report for dump precision comparison."""
    lines = []
    lines.append("# Dump 数据精度对比报告 (Ascend vs H20)\n")

    # ── Configuration ──
    lines.append("## 1. 配置信息\n")
    lines.append("| 项目 | 值 |")
    lines.append("|---|---|")
    lines.append(f"| Baseline (H20) 目录 | `{baseline_dir}` |")
    lines.append(f"| Test (Ascend) 目录 | `{test_dir}` |")
    lines.append(f"| 比对文件数 | {len(results)} |")

    layer_set = sorted(set(r.layer_idx for r in results))
    lines.append(f"| 层范围 | {layer_set[0]} ~ {layer_set[-1]} ({len(layer_set)} 层) |")
    lines.append("")

    # ── Full comparison table by dump order ──
    lines.append("## 2. 按 Dump 顺序的完整比对结果\n")

    current_layer = None
    seq_num = 0
    for r in results:
        seq_num += 1
        # Print layer header when layer changes
        if r.layer_idx != current_layer:
            current_layer = r.layer_idx
            has_moe = any(rr.layer_idx == r.layer_idx and "experts" in rr.dump_name for rr in results)
            layer_type = "MoE" if has_moe else "Dense"

            lines.append(f"\n### Layer {r.layer_idx} ({layer_type})\n")
            lines.append("| # | Dump Point | Stage | Shape | SNR(dB) | mean_ulp | within_1ulp% | within_2ulp% | mean_abs_err | max_abs_err |")
            lines.append("|---|---|---|---|---|---|---|---|---|---|")

        shape_str = "\u00d7".join(str(s) for s in r.baseline_shape)
        lines.append(
            f"| {seq_num} | `{r.dump_name}` | {r.stage} | "
            f"{shape_str} | "
            f"{_fmt(r.snr_db, '.2f')} | "
            f"{_fmt(r.mean_ulp, '.4f')} | "
            f"{_fmt(r.within_1ulp_pct, '.2f')}% | "
            f"{_fmt(r.within_2ulp_pct, '.2f')}% | "
            f"{r.mean_abs_err:.2e} | "
            f"{r.max_abs_err:.2e} |"
        )
    lines.append("")

    # ── Summary by stage ──
    lines.append("## 3. 按功能阶段汇总\n")
    lines.append("| 阶段 | 文件数 | 平均 SNR(dB) | 最小 SNR(dB) | 平均 within_1ulp% | 最小 within_1ulp% | 平均 mean_ulp |")
    lines.append("|---|---|---|---|---|---|---|")

    seen_stages = []
    for r in results:
        if r.stage not in seen_stages:
            seen_stages.append(r.stage)

    for stage in seen_stages:
        stage_results = [r for r in results if r.stage == stage]
        finite_snrs = [r.snr_db for r in stage_results if math.isfinite(r.snr_db)]

        n = len(stage_results)
        avg_snr = sum(finite_snrs) / len(finite_snrs) if finite_snrs else float('nan')
        min_snr = min(finite_snrs) if finite_snrs else float('nan')
        avg_w1 = sum(r.within_1ulp_pct for r in stage_results) / n
        min_w1 = min(r.within_1ulp_pct for r in stage_results)
        avg_ulp = sum(r.mean_ulp for r in stage_results) / n

        lines.append(
            f"| {stage} | {n} | "
            f"{_fmt(avg_snr, '.2f')} | {_fmt(min_snr, '.2f')} | "
            f"{_fmt(avg_w1, '.2f')}% | {_fmt(min_w1, '.2f')}% | "
            f"{_fmt(avg_ulp, '.4f')} |"
        )
    lines.append("")

    # ── Summary by layer ──
    lines.append("## 4. 按层汇总\n")
    lines.append("| 层 | 文件数 | 平均 SNR(dB) | 最小 SNR(dB) | 平均 within_1ulp% | 最小 within_1ulp% | 平均 mean_ulp |")
    lines.append("|---|---|---|---|---|---|---|")

    for layer_idx in layer_set:
        layer_results = [r for r in results if r.layer_idx == layer_idx]
        finite_snrs = [r.snr_db for r in layer_results if math.isfinite(r.snr_db)]

        n = len(layer_results)
        avg_snr = sum(finite_snrs) / len(finite_snrs) if finite_snrs else float('nan')
        min_snr = min(finite_snrs) if finite_snrs else float('nan')
        avg_w1 = sum(r.within_1ulp_pct for r in layer_results) / n
        min_w1 = min(r.within_1ulp_pct for r in layer_results)
        avg_ulp = sum(r.mean_ulp for r in layer_results) / n

        lines.append(
            f"| {layer_idx} | {n} | "
            f"{_fmt(avg_snr, '.2f')} | {_fmt(min_snr, '.2f')} | "
            f"{_fmt(avg_w1, '.2f')}% | {_fmt(min_w1, '.2f')}% | "
            f"{_fmt(avg_ulp, '.4f')} |"
        )
    lines.append("")

    # ── Quality Assessment ──
    lines.append("## 5. 质量评估\n")

    finite_snrs = [r.snr_db for r in results if math.isfinite(r.snr_db)]
    overall_avg_snr = sum(finite_snrs) / len(finite_snrs) if finite_snrs else float('nan')
    overall_avg_w1 = sum(r.within_1ulp_pct for r in results) / len(results)
    overall_avg_ulp = sum(r.mean_ulp for r in results) / len(results)

    lines.append("### 总体指标\n")
    lines.append("| 指标 | 值 | 评估 |")
    lines.append("|---|---|---|")

    if math.isfinite(overall_avg_snr):
        if overall_avg_snr >= 60:
            snr_status = "Excellent"
        elif overall_avg_snr >= 50:
            snr_status = "Good"
        else:
            snr_status = "Needs Investigation"
    else:
        snr_status = "Perfect"
    lines.append(f"| 平均 SNR | {_fmt(overall_avg_snr, '.2f')} dB | {snr_status} |")

    if overall_avg_w1 >= 99:
        ulp_status = "Excellent"
    elif overall_avg_w1 >= 95:
        ulp_status = "Good"
    else:
        ulp_status = "Needs Investigation"
    lines.append(f"| 平均 within_1ulp | {_fmt(overall_avg_w1, '.2f')}% | {ulp_status} |")
    lines.append(f"| 平均 mean_ulp | {_fmt(overall_avg_ulp, '.4f')} | |")
    lines.append("")

    lines.append("### 各阶段评估\n")
    for stage in seen_stages:
        stage_results = [r for r in results if r.stage == stage]
        finite_snrs_s = [r.snr_db for r in stage_results if math.isfinite(r.snr_db)]
        if finite_snrs_s:
            avg_s = sum(finite_snrs_s) / len(finite_snrs_s)
            if avg_s >= 60:
                status = "Excellent"
            elif avg_s >= 50:
                status = "Good"
            else:
                status = "Needs investigation"
            lines.append(f"- **{stage}**: {status} (avg SNR = {_fmt(avg_s, '.2f')} dB)")
        else:
            lines.append(f"- **{stage}**: Perfect (no difference)")
    lines.append("")

    # ── Anomaly detection ──
    lines.append("## 6. 异常点检测\n")
    lines.append("以下 dump 点的 SNR < 50 dB 或 within_1ulp < 95%：\n")
    lines.append("| Layer | Dump Point | SNR(dB) | within_1ulp% | mean_ulp |")
    lines.append("|---|---|---|---|---|")

    anomaly_count = 0
    for r in results:
        is_anomaly = False
        if math.isfinite(r.snr_db) and r.snr_db < 50:
            is_anomaly = True
        if r.within_1ulp_pct < 95:
            is_anomaly = True
        if is_anomaly:
            anomaly_count += 1
            lines.append(
                f"| {r.layer_idx} | `{r.dump_name}` | "
                f"{_fmt(r.snr_db, '.2f')} | {_fmt(r.within_1ulp_pct, '.2f')}% | "
                f"{_fmt(r.mean_ulp, '.4f')} |"
            )

    if anomaly_count == 0:
        lines.append("| — | 无异常点 | — | — | — |")
    lines.append(f"\n共 {anomaly_count} 个异常点（占比 {anomaly_count/len(results)*100:.1f}%）")
    lines.append("")

    return "\n".join(lines)


# ===========================================================================
# Excel report generation
# ===========================================================================

def generate_excel_report(
    results: List[DumpMetrics],
    output_path: str,
) -> str:
    """Generate an Excel report following the precision_result.xlsx template.

    Template format (single sheet "precision_data"):
      Row 1: | (empty) | (empty) | H20 (merged C-E) | 950 (merged F-H) | error/diff (merged I-Q) |
      Row 2: | (layer label) | layer | mean | min | max | mean | min | max |
              mean_abs_error | max_abs_error | min_abs_error |
              mean_rel_error | max_rel_error | min_rel_error |
              mean_ulp | within_1ulp% | within_2ulp% |
      Row 3+: data rows grouped by layer, column A has "layerN" for first row of each layer group.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "precision_data"

    # ── Styles ──
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    layer_font = Font(bold=True, size=11, color="1F4E79")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ── Row 1: Group headers (merged) ──
    ws.merge_cells("C1:E1")
    ws["C1"] = "H20"
    ws["C1"].font = header_font
    ws["C1"].alignment = center_align
    ws["C1"].fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")

    ws.merge_cells("F1:H1")
    ws["F1"] = "950"
    ws["F1"].font = header_font
    ws["F1"].alignment = center_align
    ws["F1"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    ws.merge_cells("I1:Q1")
    ws["I1"] = "error/diff"
    ws["I1"].font = header_font
    ws["I1"].alignment = center_align
    ws["I1"].fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

    # ── Row 2: Column headers ──
    col_headers = [
        "",           # A: layer group label
        "layer",      # B: dump point name
        "mean",       # C: H20 mean
        "min",        # D: H20 min
        "max",        # E: H20 max
        "mean",       # F: 950 mean
        "min",        # G: 950 min
        "max",        # H: 950 max
        "mean_abs_error",   # I
        "max_abs_error",    # J
        "min_abs_error",    # K
        "mean_rel_error",   # L
        "max_rel_error",    # M
        "min_rel_error",    # N
        "mean_ulp",         # O
        "within_1ulp%",     # P
        "within_2ulp%",     # Q
    ]
    for col_idx, header in enumerate(col_headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # ── Data rows ──
    row_num = 3
    current_layer = None
    layer_start_row = None

    for r in results:
        # Write layer label in column A when layer changes
        if r.layer_idx != current_layer:
            # Merge previous layer's A column cells if applicable
            if current_layer is not None and layer_start_row is not None:
                if row_num - 1 > layer_start_row:
                    ws.merge_cells(
                        start_row=layer_start_row, start_column=1,
                        end_row=row_num - 1, end_column=1,
                    )
                cell_a = ws.cell(row=layer_start_row, column=1)
                cell_a.alignment = Alignment(horizontal="center", vertical="center")

            current_layer = r.layer_idx
            layer_start_row = row_num
            ws.cell(row=row_num, column=1, value=f"layer{r.layer_idx}").font = layer_font

        # Column B: dump point name
        ws.cell(row=row_num, column=2, value=r.dump_name)

        # Columns C-E: H20 (baseline) stats
        ws.cell(row=row_num, column=3, value=r.baseline_mean)
        ws.cell(row=row_num, column=4, value=r.baseline_min)
        ws.cell(row=row_num, column=5, value=r.baseline_max)

        # Columns F-H: 950 (Ascend/test) stats
        ws.cell(row=row_num, column=6, value=r.test_mean)
        ws.cell(row=row_num, column=7, value=r.test_min)
        ws.cell(row=row_num, column=8, value=r.test_max)

        # Columns I-K: abs error
        ws.cell(row=row_num, column=9, value=r.mean_abs_err)
        ws.cell(row=row_num, column=10, value=r.max_abs_err)
        ws.cell(row=row_num, column=11, value=r.min_abs_err)

        # Columns L-N: rel error
        ws.cell(row=row_num, column=12, value=r.mean_rel_err)
        ws.cell(row=row_num, column=13, value=r.max_rel_err)
        ws.cell(row=row_num, column=14, value=r.min_rel_err)

        # Columns O-Q: ULP metrics
        ws.cell(row=row_num, column=15, value=r.mean_ulp)
        ws.cell(row=row_num, column=16, value=r.within_1ulp_pct)
        ws.cell(row=row_num, column=17, value=r.within_2ulp_pct)

        # Apply borders to data cells
        for col_idx in range(1, 18):
            ws.cell(row=row_num, column=col_idx).border = thin_border

        row_num += 1

    # Merge the last layer's A column
    if current_layer is not None and layer_start_row is not None:
        if row_num - 1 > layer_start_row:
            ws.merge_cells(
                start_row=layer_start_row, start_column=1,
                end_row=row_num - 1, end_column=1,
            )
        cell_a = ws.cell(row=layer_start_row, column=1)
        cell_a.alignment = Alignment(horizontal="center", vertical="center")

    # ── Column widths ──
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 40
    for col_letter in ["C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col_letter].width = 14
    for col_letter in ["I", "J", "K", "L", "M", "N"]:
        ws.column_dimensions[col_letter].width = 16
    for col_letter in ["O", "P", "Q"]:
        ws.column_dimensions[col_letter].width = 14

    # ── Number format for data cells ──
    for row in ws.iter_rows(min_row=3, max_row=row_num - 1, min_col=3, max_col=17):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = '0.000000'

    # ── Freeze panes ──
    ws.freeze_panes = "C3"

    # Save
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path


# ===========================================================================
# Chart generation
# ===========================================================================

def generate_charts(results: List[DumpMetrics], output_dir: str) -> list:
    """Generate precision analysis charts."""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib.colors import Normalize
    except ImportError:
        print("WARNING: matplotlib not available, skipping chart generation")
        return []

    os.makedirs(output_dir, exist_ok=True)
    saved_files = []

    layer_set = sorted(set(r.layer_idx for r in results))

    # Stage colors for grouping
    stage_colors = {
        "Layer Input": "#1f77b4",
        "Input LayerNorm": "#ff7f0e",
        "Attention": "#2ca02c",
        "Attention Output": "#98df8a",
        "Attention+Residual": "#d62728",
        "Post-Attn LayerNorm": "#ff9896",
        "Dense FFN": "#9467bd",
        "MoE Entry": "#8c564b",
        "MoE Routing": "#e377c2",
        "MoE Fusion": "#7f7f7f",
        "MoE Experts": "#bcbd22",
        "MLP+Residual": "#17becf",
        "Unknown": "#aaaaaa",
    }

    # ── Chart 1: SNR by layer (key dump points) ──
    fig, ax = plt.subplots(figsize=(16, 7))

    key_dumps = [
        "after_input_layernorm",
        "after_attn_add_residual",
        "after_post_attention_layernorm",
        "after_router",
        "after_mlp_layer",
    ]
    key_labels = {
        "after_input_layernorm": "Input LayerNorm",
        "after_attn_add_residual": "Attn+Residual",
        "after_post_attention_layernorm": "Post-Attn LN",
        "after_router": "Router Output",
        "after_mlp_layer": "MLP+Residual",
    }
    key_colors = ["#ff7f0e", "#d62728", "#ff9896", "#8c564b", "#17becf"]

    for i, dump_name in enumerate(key_dumps):
        x_vals = []
        y_vals = []
        for layer_idx in layer_set:
            matched = [r for r in results if r.layer_idx == layer_idx and r.dump_name == dump_name]
            if matched:
                r = matched[0]
                x_vals.append(layer_idx)
                y_vals.append(r.snr_db if math.isfinite(r.snr_db) else 100.0)
        if x_vals:
            ax.plot(x_vals, y_vals, marker='o', markersize=3, linewidth=1.5,
                    color=key_colors[i], label=key_labels[dump_name])

    ax.set_xlabel("Layer Index", fontsize=12)
    ax.set_ylabel("SNR (dB)", fontsize=12)
    ax.set_title("Key Dump Points: SNR by Layer (Higher = Better)", fontsize=14)
    ax.legend(fontsize=10)
    ax.grid(True, alpha=0.3)
    ax.axhline(y=50, color="green", linestyle="--", alpha=0.5, label="Good (50 dB)")
    ax.axhline(y=60, color="blue", linestyle="--", alpha=0.3, label="Excellent (60 dB)")
    fig.tight_layout()
    path = os.path.join(output_dir, "snr_by_layer.png")
    fig.savefig(path, dpi=150)
    plt.close(fig)
    saved_files.append(path)
    print(f"  Saved: {path}")

    # ── Chart 2: within_1ulp by layer (key dump points) ──
    fig, ax = plt.subplots(figsize=(16, 7))

    for i, dump_name in enumerate(key_dumps):
        x_vals = []
        y_vals = []
        for layer_idx in layer_set:
            matched = [r for r in results if r.layer_idx == layer_idx and r.dump_name == dump_name]
            if matched:
                r = matched[0]
                x_vals.append(layer_idx)
                y_vals.append(r.within_1ulp_pct)
        if x_vals:
            ax.plot(x_vals, y_vals, marker='o', markersize=3, linewidth=1.5,
                    color=key_colors[i], label=key_labels[dump_name])

    ax.set_xlabel("Layer Index", fontsize=12)
    ax.set_ylabel("within_1ulp (%)", fontsize=12)
    ax.set_title("Key Dump Points: within_1ULP by Layer (Higher = Better)", fontsize=14)
    ax.legend(fontsize=10)
    ax.grid(True, alpha=0.3)
    ax.axhline(y=95, color="green", linestyle="--", alpha=0.5)
    ax.axhline(y=99, color="blue", linestyle="--", alpha=0.3)
    all_w1 = [r.within_1ulp_pct for r in results]
    ax.set_ylim(bottom=max(0, min(all_w1) - 5), top=105)
    fig.tight_layout()
    path = os.path.join(output_dir, "within_1ulp_by_layer.png")
    fig.savefig(path, dpi=150)
    plt.close(fig)
    saved_files.append(path)
    print(f"  Saved: {path}")

    # ── Chart 3: SNR Heatmap ──
    moe_layers = [l for l in layer_set if l >= 1]
    heatmap_dumps = MOE_LAYER_DUMPS if moe_layers else DENSE_LAYER_DUMPS

    heatmap_layers = [l for l in layer_set if l >= 0]
    heatmap_data = np.full((len(heatmap_layers), len(heatmap_dumps)), np.nan)

    for i, layer_idx in enumerate(heatmap_layers):
        for j, dump_name in enumerate(heatmap_dumps):
            matched = [r for r in results if r.layer_idx == layer_idx and r.dump_name == dump_name]
            if matched:
                val = matched[0].snr_db
                heatmap_data[i, j] = val if math.isfinite(val) else 100.0

    fig, ax = plt.subplots(figsize=(20, max(10, len(heatmap_layers) * 0.35)))
    heatmap_display = np.nan_to_num(heatmap_data, nan=0.0)
    vmax = min(100.0, max(70.0, np.nanmax(heatmap_display)))

    im = ax.imshow(
        heatmap_display, aspect="auto", cmap="RdYlGn",
        norm=Normalize(vmin=0, vmax=vmax),
    )

    ax.set_xticks(range(len(heatmap_dumps)))
    short_names = [d.replace("grouped_topk_", "gt_").replace("after_", "")
                   for d in heatmap_dumps]
    ax.set_xticklabels(short_names, fontsize=7, rotation=45, ha='right')

    ax.set_yticks(range(len(heatmap_layers)))
    ax.set_yticklabels([str(l) for l in heatmap_layers], fontsize=8)

    ax.set_xlabel("Dump Point", fontsize=12)
    ax.set_ylabel("Layer Index", fontsize=12)
    ax.set_title("SNR Heatmap (dB) - All Layers x All Dump Points", fontsize=14)

    fig.colorbar(im, ax=ax, label="SNR (dB)", shrink=0.8)
    fig.tight_layout()
    path = os.path.join(output_dir, "snr_heatmap.png")
    fig.savefig(path, dpi=150)
    plt.close(fig)
    saved_files.append(path)
    print(f"  Saved: {path}")

    # ── Chart 4: Error accumulation (full dump order) ──
    fig, ax = plt.subplots(figsize=(20, 7))

    snr_values = []
    x_positions = []
    colors = []
    layer_boundaries = []
    prev_layer = None

    for i, r in enumerate(results):
        snr_val = r.snr_db if math.isfinite(r.snr_db) else 100.0
        snr_values.append(snr_val)
        x_positions.append(i)
        colors.append(stage_colors.get(r.stage, "#aaaaaa"))

        if r.layer_idx != prev_layer:
            if prev_layer is not None:
                layer_boundaries.append(i)
            prev_layer = r.layer_idx

    ax.plot(x_positions, snr_values, linewidth=0.8, alpha=0.7, color="#1f77b4")
    ax.scatter(x_positions, snr_values, c=colors, s=5, alpha=0.8, zorder=5)

    for boundary in layer_boundaries:
        ax.axvline(x=boundary, color="gray", linestyle=":", alpha=0.3)

    ax.set_xlabel("Dump Sequence Index (Execution Order)", fontsize=12)
    ax.set_ylabel("SNR (dB)", fontsize=12)
    ax.set_title("Error Accumulation: SNR Through All Layers and Dump Points", fontsize=14)
    ax.grid(True, alpha=0.3)
    ax.axhline(y=50, color="green", linestyle="--", alpha=0.5)
    ax.axhline(y=60, color="blue", linestyle="--", alpha=0.3)

    from matplotlib.patches import Patch
    legend_elements = [Patch(facecolor=c, label=s) for s, c in stage_colors.items()
                       if any(r.stage == s for r in results)]
    ax.legend(handles=legend_elements, fontsize=7, loc='lower left', ncol=3)

    fig.tight_layout()
    path = os.path.join(output_dir, "error_accumulation.png")
    fig.savefig(path, dpi=150)
    plt.close(fig)
    saved_files.append(path)
    print(f"  Saved: {path}")

    return saved_files


# ===========================================================================
# Main
# ===========================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Compare dump data precision between Ascend and H20 (ULP + SNR)"
    )
    parser.add_argument(
        "--baseline-dir", type=str, default="./dump_data_h20",
        help="Directory with baseline dump data (default: ./dump_data_h20)",
    )
    parser.add_argument(
        "--test-dir", type=str, default="./dump_data_ascend",
        help="Directory with test dump data (default: ./dump_data_ascend)",
    )
    parser.add_argument(
        "--output-report", "-o", type=str, default=None,
        help="Path to save markdown report (default: stdout only)",
    )
    parser.add_argument(
        "--output-charts-dir", type=str, default=None,
        help="Directory to save charts (default: same as report dir)",
    )
    parser.add_argument(
        "--output-excel", type=str, default=None,
        help="Path to save Excel report (default: auto based on report path)",
    )
    args = parser.parse_args()

    print(f"Dump data precision comparison:")
    print(f"  Baseline (H20):    {args.baseline_dir}")
    print(f"  Test (Ascend):     {args.test_dir}")
    print()

    # Run comparison
    results = compare_dumps(args.baseline_dir, args.test_dir)

    if not results:
        print("ERROR: No dump files could be compared!")
        return

    # Generate markdown report
    md_report = generate_markdown_report(results, args.baseline_dir, args.test_dir)

    if args.output_report:
        os.makedirs(os.path.dirname(args.output_report) or ".", exist_ok=True)
        with open(args.output_report, "w") as f:
            f.write(md_report)
        print(f"\nReport saved to {args.output_report}")

    print(md_report)

    # Generate charts
    charts_dir = args.output_charts_dir
    if not charts_dir and args.output_report:
        charts_dir = os.path.join(
            os.path.dirname(args.output_report) or ".", "dump_charts"
        )

    if charts_dir:
        print(f"\nGenerating charts to {charts_dir}...")
        saved = generate_charts(results, charts_dir)
        print(f"  Generated {len(saved)} charts")

    # Generate Excel report
    excel_path = args.output_excel
    if not excel_path and args.output_report:
        base = os.path.splitext(args.output_report)[0]
        excel_path = base + ".xlsx"

    if excel_path:
        print(f"\nGenerating Excel report...")
        saved_path = generate_excel_report(results, excel_path)
        print(f"  Saved: {saved_path}")


if __name__ == "__main__":
    main()
